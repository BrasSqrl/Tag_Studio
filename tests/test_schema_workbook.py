from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from tag_studio.defaults import DEFAULT_OUTCOME_TAXONOMY, DEFAULT_SCORING_RUBRIC, DEFAULT_SECTIONS, DEFAULT_TAGS
from tag_studio.schema_workbook import TAG_SHEET, create_tag_setup_workbook, import_tag_setup_workbook


def test_tag_setup_workbook_round_trips_defaults() -> None:
    workbook = create_tag_setup_workbook(
        DEFAULT_SECTIONS,
        DEFAULT_TAGS,
        DEFAULT_OUTCOME_TAXONOMY,
        [record.model_dump() for record in DEFAULT_SCORING_RUBRIC],
    )
    result = import_tag_setup_workbook(
        workbook,
        DEFAULT_SECTIONS,
        DEFAULT_TAGS,
        update_sections=True,
        update_tags=True,
        update_outcomes=True,
        update_scoring=True,
    )

    assert result.section_count == len(DEFAULT_SECTIONS)
    assert result.tag_count == len(DEFAULT_TAGS)
    assert len(result.outcomes or []) == len(DEFAULT_OUTCOME_TAXONOMY)
    assert len(result.scoring or []) == len(DEFAULT_SCORING_RUBRIC)


def test_tag_setup_import_rejects_invalid_data_type() -> None:
    workbook = create_tag_setup_workbook(DEFAULT_SECTIONS, DEFAULT_TAGS)
    wb = load_workbook(BytesIO(workbook))
    ws = wb[TAG_SHEET]
    ws["D2"] = "not_a_type"
    output = BytesIO()
    wb.save(output)

    with pytest.raises(ValueError, match="data_type"):
        import_tag_setup_workbook(
            output.getvalue(),
            DEFAULT_SECTIONS,
            DEFAULT_TAGS,
            update_sections=False,
            update_tags=True,
        )
