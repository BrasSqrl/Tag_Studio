from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .defaults import DEFAULT_FACILITY_TYPES, DEFAULT_MEMO_TYPES, SCHEMA_VERSION
from .models import SectionDefinition, TagDefinition

SECTION_SHEET = "Standard Memo Sections"
TAG_SHEET = "Credit Tags"
FACILITY_TYPES_SHEET = "Facility Types"
OUTCOME_SHEET = "Outcome Taxonomy"
SCORING_SHEET = "Scoring Rubric"
EVIDENCE_RULES_SHEET = "Evidence Rules"
EXPORT_MAPPING_SHEET = "Export Mapping"
INSTRUCTIONS_SHEET = "Instructions"
ALLOWED_VALUES_SHEET = "Allowed Values"

SECTION_COLUMNS = [
    "section_id",
    "display_name",
    "description",
    "required",
    "memo_types",
    "facility_types",
    "expected_tag_ids",
    "evidence_required",
    "aliases",
    "display_order",
]

TAG_COLUMNS = [
    "tag_id",
    "label",
    "category",
    "data_type",
    "allowed_values",
    "required",
    "evidence_required",
    "material",
    "allowed_scopes",
    "default_scope",
    "facility_required",
    "scoring_use",
    "export_use",
    "help_text",
]

OUTCOME_COLUMNS = ["label", "severity_rank"]
SCORING_COLUMNS = [
    "score_name",
    "component_tag_id",
    "weight",
    "directionality",
    "min_value",
    "max_value",
    "required_evidence",
    "memo_types",
    "facility_types",
    "active",
    "version",
]

DATA_TYPES = ["text", "long_text", "enum", "multi_select", "number", "boolean"]
EXPORT_USES = ["section", "memo", "both", "none"]
TAG_SCOPES = ["memo", "borrower", "facility", "section", "outcome"]


@dataclass
class TagSetupImportResult:
    sections: list[SectionDefinition] | None = None
    tags: list[TagDefinition] | None = None
    outcomes: list[dict[str, Any]] | None = None
    scoring: list[dict[str, Any]] | None = None
    warnings: list[str] = field(default_factory=list)

    @property
    def section_count(self) -> int:
        return len(self.sections or [])

    @property
    def tag_count(self) -> int:
        return len(self.tags or [])


def _slugify(value: str) -> str:
    safe = "".join(ch.lower() if ch.isalnum() else "_" for ch in value.strip())
    while "__" in safe:
        safe = safe.replace("__", "_")
    return safe.strip("_") or "item"


def _join(values: list[str]) -> str:
    return "; ".join(values)


def _split_list(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    text = text.replace("\n", ";")
    return [item.strip() for item in text.split(";") if item.strip()]


def _parse_bool(value: Any, field_name: str, row_number: int, default: bool = False) -> bool:
    if value is None or str(value).strip() == "":
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in {"true", "yes", "y", "1", "required"}:
        return True
    if normalized in {"false", "no", "n", "0", "optional"}:
        return False
    raise ValueError(f"Row {row_number}: {field_name} must be TRUE or FALSE.")


def _parse_int(value: Any, field_name: str, row_number: int, default: int = 100) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_number}: {field_name} must be a whole number.") from exc


def _style_sheet(ws, columns: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="102033")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_index, header in enumerate(columns, start=1):
        longest = len(header)
        for row in ws.iter_rows(min_row=2, min_col=column_index, max_col=column_index):
            value = row[0].value
            if value is not None:
                longest = max(longest, min(len(str(value)), 70))
        ws.column_dimensions[ws.cell(row=1, column=column_index).column_letter].width = min(max(longest + 2, 14), 64)


def _add_validation(ws, cell_range: str, values: list[str]) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(cell_range)


def _write_instructions(wb: Workbook) -> None:
    ws = wb.active
    ws.title = INSTRUCTIONS_SHEET
    rows = [
        ["Tag Studio Tag Setup Template", ""],
        ["Purpose", "Edit this workbook to update the standard memo sections and credit tags used by Tag Studio."],
        ["How to edit", "Change values in the Standard Memo Sections and Credit Tags tabs. Keep sheet names and header rows unchanged."],
        ["List fields", "Use semicolons for list fields such as aliases, memo_types, facility_types, expected_tag_ids, and allowed_values."],
        ["Stable IDs", "Do not change section_id or tag_id after those IDs are used in reviewed memos unless you intentionally want a new field."],
        ["Required sections", "Use required=TRUE, then limit by memo_types and facility_types only when the section is conditional."],
        ["Import behavior", "Uploading this workbook replaces the selected setup tabs in Admin Tools."],
        ["Schema version", SCHEMA_VERSION],
    ]
    for row in rows:
        ws.append(row)
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="102033")
    ws["B1"].fill = PatternFill("solid", fgColor="102033")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_allowed_values(wb: Workbook) -> None:
    ws = wb.create_sheet(ALLOWED_VALUES_SHEET)
    rows = [
        ["Field", "Allowed values / examples"],
        ["required", "TRUE; FALSE"],
        ["evidence_required", "TRUE; FALSE"],
        ["data_type", _join(DATA_TYPES)],
        ["export_use", _join(EXPORT_USES)],
        ["tag_scope", _join(TAG_SCOPES)],
        ["memo_types", _join(DEFAULT_MEMO_TYPES)],
        ["facility_types", _join(DEFAULT_FACILITY_TYPES)],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, ["Field", "Allowed values / examples"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95


def create_tag_setup_workbook(
    sections: list[SectionDefinition],
    tags: list[TagDefinition],
    outcomes: list[dict[str, Any]] | None = None,
    scoring: list[dict[str, Any]] | None = None,
) -> bytes:
    wb = Workbook()
    _write_instructions(wb)
    _write_allowed_values(wb)

    section_ws = wb.create_sheet(SECTION_SHEET)
    section_ws.append(SECTION_COLUMNS)
    for section in sorted(sections, key=lambda item: item.display_order):
        section_ws.append(
            [
                section.section_id,
                section.display_name,
                section.description,
                section.required,
                _join(section.memo_types),
                _join(section.facility_types),
                _join(section.expected_tag_ids),
                section.evidence_required,
                _join(section.aliases),
                section.display_order,
            ]
        )
    _style_sheet(section_ws, SECTION_COLUMNS)
    _add_validation(section_ws, "D2:D500", ["TRUE", "FALSE"])
    _add_validation(section_ws, "H2:H500", ["TRUE", "FALSE"])

    tag_ws = wb.create_sheet(TAG_SHEET)
    tag_ws.append(TAG_COLUMNS)
    for tag in sorted(tags, key=lambda item: (item.category, item.label)):
        tag_ws.append(
            [
                tag.tag_id,
                tag.label,
                tag.category,
                tag.data_type,
                _join(tag.allowed_values),
                tag.required,
                tag.evidence_required,
                tag.material,
                _join([str(scope) for scope in tag.allowed_scopes]),
                tag.default_scope,
                tag.facility_required,
                tag.scoring_use,
                tag.export_use,
                tag.help_text,
            ]
        )
    _style_sheet(tag_ws, TAG_COLUMNS)
    _add_validation(tag_ws, "D2:D1000", DATA_TYPES)
    _add_validation(tag_ws, "F2:F1000", ["TRUE", "FALSE"])
    _add_validation(tag_ws, "G2:G1000", ["TRUE", "FALSE"])
    _add_validation(tag_ws, "H2:H1000", ["TRUE", "FALSE"])
    _add_validation(tag_ws, "J2:J1000", TAG_SCOPES)
    _add_validation(tag_ws, "K2:K1000", ["TRUE", "FALSE"])
    _add_validation(tag_ws, "M2:M1000", EXPORT_USES)

    facility_ws = wb.create_sheet(FACILITY_TYPES_SHEET)
    facility_ws.append(["facility_type"])
    for facility_type in DEFAULT_FACILITY_TYPES:
        facility_ws.append([facility_type])
    _style_sheet(facility_ws, ["facility_type"])

    outcome_ws = wb.create_sheet(OUTCOME_SHEET)
    outcome_ws.append(OUTCOME_COLUMNS)
    for outcome in outcomes or []:
        outcome_ws.append([outcome.get("label", ""), outcome.get("severity_rank", 0)])
    _style_sheet(outcome_ws, OUTCOME_COLUMNS)

    scoring_ws = wb.create_sheet(SCORING_SHEET)
    scoring_ws.append(SCORING_COLUMNS)
    for record in scoring or []:
        scoring_ws.append(
            [
                _join(record.get(column, [])) if column in {"memo_types", "facility_types"} else record.get(column, "")
                for column in SCORING_COLUMNS
            ]
        )
    _style_sheet(scoring_ws, SCORING_COLUMNS)

    evidence_ws = wb.create_sheet(EVIDENCE_RULES_SHEET)
    evidence_ws.append(["rule", "setting"])
    evidence_ws.append(["material_tags_require_exact_evidence", "TRUE"])
    evidence_ws.append(["unresolved_text_warnings_block_export", "TRUE"])
    _style_sheet(evidence_ws, ["rule", "setting"])

    export_ws = wb.create_sheet(EXPORT_MAPPING_SHEET)
    export_ws.append(["export_name", "description"])
    export_ws.append(["normalized_dataset", "Authoritative dataset used to derive training files."])
    export_ws.append(["span_level_jsonl", "Evidence span plus tag examples."])
    export_ws.append(["section_level_jsonl", "Section context plus normalized tags."])
    export_ws.append(["memo_level_jsonl", "Whole memo synthesis and score examples."])
    _style_sheet(export_ws, ["export_name", "description"])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _worksheet_rows(wb, sheet_name: str, required_columns: list[str]) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook is missing the '{sheet_name}' sheet.")
    ws = wb[sheet_name]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    lookup = {header: index for index, header in enumerate(headers)}
    missing = [column for column in required_columns if column not in lookup]
    if missing:
        raise ValueError(f"'{sheet_name}' is missing required column(s): {', '.join(missing)}.")

    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = {column: values[lookup[column]] if lookup[column] < len(values) else None for column in required_columns}
        if any(value is not None and str(value).strip() for value in row.values()):
            row["_row_number"] = row_number
            rows.append(row)
    return rows


def _validate_unique(ids: list[str], label: str) -> None:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for item_id in ids:
        if item_id in seen:
            duplicates.add(item_id)
        seen.add(item_id)
    if duplicates:
        raise ValueError(f"Duplicate {label} ID(s): {', '.join(sorted(duplicates))}.")


def _parse_sections(wb) -> list[SectionDefinition]:
    rows = _worksheet_rows(wb, SECTION_SHEET, SECTION_COLUMNS)
    sections: list[SectionDefinition] = []
    for row in rows:
        row_number = int(row["_row_number"])
        raw_id = str(row.get("section_id") or "").strip()
        if not raw_id:
            raise ValueError(f"Row {row_number}: section_id is required.")
        section_id = _slugify(raw_id)
        display_name = str(row.get("display_name") or raw_id).strip()
        sections.append(
            SectionDefinition(
                section_id=section_id,
                display_name=display_name,
                description=str(row.get("description") or "").strip(),
                required=_parse_bool(row.get("required"), "required", row_number, default=True),
                memo_types=_split_list(row.get("memo_types")),
                facility_types=_split_list(row.get("facility_types")),
                expected_tag_ids=[_slugify(item) for item in _split_list(row.get("expected_tag_ids"))],
                evidence_required=_parse_bool(row.get("evidence_required"), "evidence_required", row_number, default=True),
                aliases=_split_list(row.get("aliases")),
                display_order=_parse_int(row.get("display_order"), "display_order", row_number),
            )
        )
    if not sections:
        raise ValueError(f"'{SECTION_SHEET}' must contain at least one section.")
    _validate_unique([section.section_id for section in sections], "section")
    return sorted(sections, key=lambda item: item.display_order)


def _parse_tags(wb) -> list[TagDefinition]:
    rows = _worksheet_rows(wb, TAG_SHEET, TAG_COLUMNS)
    tags: list[TagDefinition] = []
    for row in rows:
        row_number = int(row["_row_number"])
        raw_id = str(row.get("tag_id") or "").strip()
        if not raw_id:
            raise ValueError(f"Row {row_number}: tag_id is required.")
        tag_id = _slugify(raw_id)
        data_type = str(row.get("data_type") or "text").strip()
        if data_type not in DATA_TYPES:
            raise ValueError(f"Row {row_number}: data_type must be one of: {', '.join(DATA_TYPES)}.")
        export_use = str(row.get("export_use") or "both").strip()
        if export_use not in EXPORT_USES:
            raise ValueError(f"Row {row_number}: export_use must be one of: {', '.join(EXPORT_USES)}.")
        allowed_scopes = _split_list(row.get("allowed_scopes")) or ["section"]
        invalid_scopes = [scope for scope in allowed_scopes if scope not in TAG_SCOPES]
        if invalid_scopes:
            raise ValueError(f"Row {row_number}: allowed_scopes contains invalid value(s): {', '.join(invalid_scopes)}.")
        default_scope = str(row.get("default_scope") or allowed_scopes[0]).strip()
        if default_scope not in TAG_SCOPES:
            raise ValueError(f"Row {row_number}: default_scope must be one of: {', '.join(TAG_SCOPES)}.")
        if default_scope not in allowed_scopes:
            allowed_scopes.append(default_scope)
        tags.append(
            TagDefinition(
                tag_id=tag_id,
                label=str(row.get("label") or raw_id).strip(),
                category=str(row.get("category") or "General").strip(),
                data_type=data_type,  # type: ignore[arg-type]
                allowed_values=_split_list(row.get("allowed_values")),
                required=_parse_bool(row.get("required"), "required", row_number),
                evidence_required=_parse_bool(row.get("evidence_required"), "evidence_required", row_number),
                material=_parse_bool(row.get("material"), "material", row_number),
                allowed_scopes=allowed_scopes,  # type: ignore[arg-type]
                default_scope=default_scope,  # type: ignore[arg-type]
                facility_required=_parse_bool(row.get("facility_required"), "facility_required", row_number),
                scoring_use=str(row.get("scoring_use") or "").strip(),
                export_use=export_use,  # type: ignore[arg-type]
                help_text=str(row.get("help_text") or "").strip(),
            )
        )
    if not tags:
        raise ValueError(f"'{TAG_SHEET}' must contain at least one tag.")
    _validate_unique([tag.tag_id for tag in tags], "tag")
    return tags


def _parse_outcomes(wb) -> list[dict[str, Any]] | None:
    if OUTCOME_SHEET not in wb.sheetnames:
        return None
    rows = _worksheet_rows(wb, OUTCOME_SHEET, OUTCOME_COLUMNS)
    outcomes = []
    for row in rows:
        if not row.get("label"):
            continue
        outcomes.append({"label": str(row.get("label")).strip(), "severity_rank": _parse_int(row.get("severity_rank"), "severity_rank", int(row["_row_number"]), 0)})
    return outcomes


def _parse_scoring(wb) -> list[dict[str, Any]] | None:
    if SCORING_SHEET not in wb.sheetnames:
        return None
    rows = _worksheet_rows(wb, SCORING_SHEET, SCORING_COLUMNS)
    scoring = []
    for row in rows:
        if not row.get("score_name") or not row.get("component_tag_id"):
            continue
        row_number = int(row["_row_number"])
        scoring.append(
            {
                "score_name": str(row.get("score_name")).strip(),
                "component_tag_id": str(row.get("component_tag_id")).strip(),
                "weight": float(row.get("weight") or 0),
                "directionality": str(row.get("directionality") or "higher_is_better"),
                "min_value": float(row.get("min_value") or 0),
                "max_value": float(row.get("max_value") or 100),
                "required_evidence": _parse_bool(row.get("required_evidence"), "required_evidence", row_number, True),
                "memo_types": _split_list(row.get("memo_types")),
                "facility_types": _split_list(row.get("facility_types")),
                "active": _parse_bool(row.get("active"), "active", row_number, True),
                "version": str(row.get("version") or SCHEMA_VERSION),
            }
        )
    return scoring


def import_tag_setup_workbook(
    workbook_bytes: bytes,
    current_sections: list[SectionDefinition],
    current_tags: list[TagDefinition],
    *,
    update_sections: bool,
    update_tags: bool,
    update_outcomes: bool = False,
    update_scoring: bool = False,
) -> TagSetupImportResult:
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    result = TagSetupImportResult()

    if update_sections:
        result.sections = _parse_sections(wb)
    if update_tags:
        result.tags = _parse_tags(wb)
    if update_outcomes:
        result.outcomes = _parse_outcomes(wb)
    if update_scoring:
        result.scoring = _parse_scoring(wb)

    final_sections = result.sections if result.sections is not None else current_sections
    final_tags = result.tags if result.tags is not None else current_tags
    tag_ids = {tag.tag_id for tag in final_tags}

    for section in final_sections:
        unknown_tags = [tag_id for tag_id in section.expected_tag_ids if tag_id not in tag_ids]
        if unknown_tags:
            result.warnings.append(
                f"{section.display_name} references expected tag ID(s) not found in Credit Tags: {', '.join(unknown_tags)}."
            )

    if result.sections is not None:
        removed_sections = sorted({section.section_id for section in current_sections} - {section.section_id for section in result.sections})
        if removed_sections:
            result.warnings.append(
                "Removed standard section ID(s): "
                + ", ".join(removed_sections)
                + ". Existing reviewed memo records will still store their prior section IDs."
            )
    if result.tags is not None:
        removed_tags = sorted({tag.tag_id for tag in current_tags} - {tag.tag_id for tag in result.tags})
        if removed_tags:
            result.warnings.append(
                "Removed credit tag ID(s): "
                + ", ".join(removed_tags)
                + ". Existing reviewed memo records will still store their prior tag IDs."
            )

    return result
