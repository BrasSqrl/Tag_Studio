from __future__ import annotations

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd

from .defaults import SCHEMA_VERSION
from .storage import (
    active_schema_hash,
    list_memo_ids,
    load_audit_events,
    load_evidence,
    load_facilities,
    load_foreseeability_assessments,
    load_memo_record,
    load_outcome_events,
    load_outcome_summaries,
    load_review,
    load_sections,
    load_table_metrics,
    load_tags,
    memo_dir,
    read_json,
    sync_path_to_remote,
)


def _now_stamp() -> str:
    return datetime.now(UTC).strftime("%Y%m%d_%H%M%S")


REVIEWER_IDENTITY_KEYS = {"reviewer", "assigned_to", "adjudicator", "tagger"}
TRAINING_IDENTITY_KEYS = {
    *REVIEWER_IDENTITY_KEYS,
    "memo_id",
    "facility_id",
    "section_id",
    "tag_record_id",
    "evidence_id",
    "outcome_summary_id",
    "outcome_event_id",
    "foreseeability_id",
    "customer_id",
    "borrower_name_or_hash",
    "borrower_id",
    "source_hash",
    "source_document_hash",
    "source_file_name",
}


def _strip_keys(data: Any, keys: set[str]) -> Any:
    if isinstance(data, list):
        return [_strip_keys(item, keys) for item in data]
    if isinstance(data, dict):
        return {key: _strip_keys(value, keys) for key, value in data.items() if key not in keys}
    return data


def _safe_sheet_rows(rows: list[dict[str, Any]], *, strip_reviewer_identity: bool = True) -> list[dict[str, Any]]:
    if not strip_reviewer_identity:
        return rows
    return [_strip_keys(row, REVIEWER_IDENTITY_KEYS) for row in rows]


def _approved_memo_ids(workspace: Path, include_only_approved: bool, include_legacy_approved: bool = False) -> list[str]:
    memo_ids = list_memo_ids(workspace)
    if not include_only_approved:
        return memo_ids
    current_hash = active_schema_hash(workspace)
    approved = []
    for memo_id in memo_ids:
        memo = load_memo_record(workspace, memo_id)
        review = load_review(workspace, memo_id)
        if review.get("status") != "Approved Gold":
            continue
        if not include_legacy_approved and memo.get("schema_hash") and memo.get("schema_hash") != current_hash:
            continue
        approved.append(memo_id)
    return approved


def build_export_tables(
    workspace: Path,
    include_only_approved: bool = True,
    include_legacy_approved: bool = False,
) -> dict[str, list[dict[str, Any]]]:
    memo_ids = _approved_memo_ids(workspace, include_only_approved, include_legacy_approved)
    tables = {
        "Memos": [],
        "Sections": [],
        "Section Mapping": [],
        "Facilities": [],
        "Tags": [],
        "Evidence": [],
        "Table Metrics": [],
        "Scores": [],
        "Outcome Summaries": [],
        "Outcome Events": [],
        "Foreseeability": [],
        "Page Quality": [],
        "Extraction Warnings": [],
        "Schema Snapshot": [],
        "Audit Events": [],
        "Review Status": [],
        "Export Manifest": [],
    }
    scoring_rubric = read_json(workspace / "config" / "scoring_rubric.json", [])
    scoring_tag_ids = {str(record.get("component_tag_id")) for record in scoring_rubric}

    for memo_id in memo_ids:
        memo = load_memo_record(workspace, memo_id)
        review = load_review(workspace, memo_id)
        sections = load_sections(workspace, memo_id)
        tags = load_tags(workspace, memo_id)
        evidence = load_evidence(workspace, memo_id)
        facilities = load_facilities(workspace, memo_id)
        outcome_summaries = load_outcome_summaries(workspace, memo_id)
        outcome_events = load_outcome_events(workspace, memo_id)
        foreseeability = load_foreseeability_assessments(workspace, memo_id)
        table_metrics = load_table_metrics(workspace, memo_id)
        page_quality = read_json(memo_dir(workspace, memo_id) / "extraction" / "page_quality.json", [])
        extraction_warnings = read_json(memo_dir(workspace, memo_id) / "extraction" / "ocr_warnings.json", [])
        schema_snapshot = read_json(memo_dir(workspace, memo_id) / "schema" / "schema_snapshot.json", {})

        tables["Memos"].append(_strip_keys(memo, REVIEWER_IDENTITY_KEYS))
        tables["Review Status"].append(_strip_keys(review, REVIEWER_IDENTITY_KEYS))
        tables["Sections"].extend(sections)
        tables["Section Mapping"].extend(
            {
                "memo_id": memo_id,
                "section_id": section.get("section_id"),
                "original_header": section.get("original_header"),
                "canonical_section_id": section.get("canonical_section_id"),
                "canonical_section_name": section.get("canonical_section_name"),
                "page_start": section.get("page_start"),
                "page_end": section.get("page_end"),
                "reviewer_confirmed": section.get("reviewer_confirmed"),
            }
            for section in sections
        )
        tables["Facilities"].extend(facilities)
        tables["Tags"].extend(_safe_sheet_rows(tags))
        tables["Evidence"].extend(evidence)
        tables["Table Metrics"].extend(table_metrics)
        tables["Scores"].extend(
            _safe_sheet_rows([tag for tag in tags if tag.get("tag_id") in scoring_tag_ids or "score" in str(tag.get("tag_id", ""))])
        )
        tables["Outcome Summaries"].extend(outcome_summaries)
        tables["Outcome Events"].extend(outcome_events)
        tables["Foreseeability"].extend(foreseeability)
        tables["Page Quality"].extend(page_quality)
        tables["Extraction Warnings"].extend(extraction_warnings)
        if schema_snapshot:
            tables["Schema Snapshot"].append(
                {
                    "memo_id": memo_id,
                    "schema_version": schema_snapshot.get("schema_version"),
                    "schema_hash": schema_snapshot.get("schema_hash"),
                    "created_at": schema_snapshot.get("created_at"),
                    "section_count": len(schema_snapshot.get("sections", [])),
                    "tag_count": len(schema_snapshot.get("tags", [])),
                    "outcome_event_type_count": len(schema_snapshot.get("outcome_taxonomy", [])),
                    "scoring_rule_count": len(schema_snapshot.get("scoring_rubric", [])),
                }
            )
        tables["Audit Events"].extend(_safe_sheet_rows(load_audit_events(workspace, memo_id)))

    tables["Export Manifest"].append(
        {
            "schema_version": SCHEMA_VERSION,
            "schema_hash": active_schema_hash(workspace),
            "created_at": datetime.now(UTC).isoformat(),
            "include_only_approved": include_only_approved,
            "include_legacy_approved": include_legacy_approved,
            "memo_count": len(memo_ids),
            "memo_ids": ", ".join(memo_ids),
        }
    )
    return tables


def export_excel(workspace: Path, include_only_approved: bool = True, include_legacy_approved: bool = False) -> Path:
    export_dir = workspace / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"tag_studio_export_{_now_stamp()}.xlsx"
    tables = build_export_tables(
        workspace,
        include_only_approved=include_only_approved,
        include_legacy_approved=include_legacy_approved,
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in tables.items():
            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame([{"note": "No records exported"}])
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 50)
    workbook.save(path)
    sync_path_to_remote(workspace, path)
    return path


def _tags_for_section(tags: list[dict[str, Any]], section_id: str) -> list[dict[str, Any]]:
    return [tag for tag in tags if tag.get("section_id") == section_id]


def _evidence_by_id(evidence: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {item.get("evidence_id"): item for item in evidence}


def _tags_for_evidence(tags: list[dict[str, Any]], evidence_id: str) -> list[dict[str, Any]]:
    return [tag for tag in tags if evidence_id in tag.get("evidence_ids", [])]


def _json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


def _record_export_ids(records: list[dict[str, Any]], id_field: str, prefix: str, memo_index: int) -> dict[str, str]:
    return {
        str(record.get(id_field)): f"{prefix}_{memo_index:06}_{index:04}"
        for index, record in enumerate(records, start=1)
        if record.get(id_field)
    }


def _training_memo(memo: dict[str, Any], export_memo_id: str) -> dict[str, Any]:
    return {
        "export_memo_id": export_memo_id,
        "memo_type": memo.get("memo_type", ""),
        "facility_type": memo.get("facility_type", ""),
        "schema_version": memo.get("schema_version", SCHEMA_VERSION),
    }


def _training_facility(facility: dict[str, Any], facility_ids: dict[str, str]) -> dict[str, Any]:
    return {
        "export_facility_id": facility_ids.get(str(facility.get("facility_id")), ""),
        "facility_type": facility.get("facility_type", ""),
        "amount": facility.get("amount", ""),
        "closing_date": facility.get("closing_date", ""),
        "status": facility.get("status", ""),
    }


def _training_section(section: dict[str, Any], section_ids: dict[str, str]) -> dict[str, Any]:
    return {
        "export_section_id": section_ids.get(str(section.get("section_id")), ""),
        "canonical_section_id": section.get("canonical_section_id"),
        "canonical_section_name": section.get("canonical_section_name"),
        "original_header": section.get("original_header"),
        "page_start": section.get("page_start"),
        "page_end": section.get("page_end"),
    }


def _training_evidence(
    evidence: dict[str, Any],
    evidence_ids: dict[str, str],
    section_ids: dict[str, str],
    facility_ids: dict[str, str],
) -> dict[str, Any]:
    return {
        "export_evidence_id": evidence_ids.get(str(evidence.get("evidence_id")), ""),
        "evidence_type": evidence.get("evidence_type", "memo_evidence"),
        "export_section_id": section_ids.get(str(evidence.get("section_id")), ""),
        "export_facility_ids": [facility_ids.get(str(item), "") for item in evidence.get("facility_ids", []) if facility_ids.get(str(item), "")],
        "page_number": evidence.get("page_number"),
        "line_start": evidence.get("line_start"),
        "line_end": evidence.get("line_end"),
        "selected_text": evidence.get("selected_text", ""),
        "source_location": evidence.get("source_location", ""),
        "evidence_role": evidence.get("evidence_role", ""),
        "citation_confidence": evidence.get("citation_confidence", ""),
    }


def _training_tag(
    tag: dict[str, Any],
    tag_ids: dict[str, str],
    facility_ids: dict[str, str],
    section_ids: dict[str, str],
    evidence_ids: dict[str, str],
) -> dict[str, Any]:
    return {
        "export_tag_id": tag_ids.get(str(tag.get("tag_record_id")), ""),
        "export_section_id": section_ids.get(str(tag.get("section_id")), ""),
        "scope": tag.get("scope", "section"),
        "export_facility_id": facility_ids.get(str(tag.get("facility_id")), "") if tag.get("facility_id") else "",
        "tag_id": tag.get("tag_id"),
        "tag_label": tag.get("tag_label"),
        "value": tag.get("value"),
        "confidence": tag.get("confidence"),
        "export_evidence_ids": [evidence_ids.get(str(item), "") for item in tag.get("evidence_ids", []) if evidence_ids.get(str(item), "")],
        "status": tag.get("status", ""),
    }


def _training_outcome_summary(
    summary: dict[str, Any],
    summary_ids: dict[str, str],
    facility_ids: dict[str, str],
    event_ids: dict[str, str],
) -> dict[str, Any]:
    return {
        "export_outcome_summary_id": summary_ids.get(str(summary.get("outcome_summary_id")), ""),
        "export_facility_id": facility_ids.get(str(summary.get("facility_id")), ""),
        "outcome_availability_state": summary.get("outcome_availability_state"),
        "seasoning_months": summary.get("seasoning_months"),
        "primary_adverse_outcome": summary.get("primary_adverse_outcome"),
        "export_primary_outcome_event_id": event_ids.get(str(summary.get("primary_outcome_event_id")), ""),
        "primary_event_date": summary.get("primary_event_date"),
        "primary_severity_rank": summary.get("primary_severity_rank"),
        "no_adverse_outcome_observed_date": summary.get("no_adverse_outcome_observed_date"),
        "source_type": summary.get("source_type"),
        "source_checked_date": summary.get("source_checked_date"),
        "source_confidence": summary.get("source_confidence"),
        "source_note": summary.get("source_note"),
    }


def _training_outcome_event(
    event: dict[str, Any],
    facility_ids: dict[str, str],
    event_ids: dict[str, str],
) -> dict[str, Any]:
    return {
        "export_outcome_event_id": event_ids.get(str(event.get("outcome_event_id")), ""),
        "export_facility_id": facility_ids.get(str(event.get("facility_id")), ""),
        "event_type": event.get("event_type"),
        "event_date": event.get("event_date"),
        "severity_rank": event.get("severity_rank"),
        "source_type": event.get("source_type"),
        "source_checked_date": event.get("source_checked_date"),
        "source_confidence": event.get("source_confidence"),
        "source_note": event.get("source_note"),
    }


def _training_foreseeability(
    assessment: dict[str, Any],
    foreseeability_ids: dict[str, str],
    facility_ids: dict[str, str],
    event_ids: dict[str, str],
    evidence_ids: dict[str, str],
) -> dict[str, Any]:
    return {
        "export_foreseeability_id": foreseeability_ids.get(str(assessment.get("foreseeability_id")), ""),
        "export_facility_id": facility_ids.get(str(assessment.get("facility_id")), ""),
        "export_outcome_event_id": event_ids.get(str(assessment.get("outcome_event_id")), ""),
        "foreseeability": assessment.get("foreseeability"),
        "export_memo_evidence_ids": [evidence_ids.get(str(item), "") for item in assessment.get("memo_evidence_ids", []) if evidence_ids.get(str(item), "")],
        "rationale": assessment.get("rationale", ""),
    }


def export_jsonl(workspace: Path, include_only_approved: bool = True, include_legacy_approved: bool = False) -> dict[str, Path]:
    export_dir = workspace / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = _now_stamp()
    span_path = export_dir / f"training_spans_{stamp}.jsonl"
    section_path = export_dir / f"training_sections_{stamp}.jsonl"
    memo_path = export_dir / f"training_memos_{stamp}.jsonl"
    outcome_path = export_dir / f"training_outcomes_{stamp}.jsonl"
    audit_path = export_dir / f"audit_records_{stamp}.jsonl"
    manifest_path = export_dir / f"training_export_manifest_{stamp}.json"

    memo_ids = _approved_memo_ids(workspace, include_only_approved, include_legacy_approved)
    manifest: dict[str, Any] = {
        "schema_version": SCHEMA_VERSION,
        "schema_hash": active_schema_hash(workspace),
        "created_at": datetime.now(UTC).isoformat(),
        "include_only_approved": include_only_approved,
        "include_legacy_approved": include_legacy_approved,
        "memo_count": len(memo_ids),
        "notes": "Training files use export-scoped IDs. This manifest preserves traceability for audit and is not model training data.",
        "memos": [],
    }
    with (
        span_path.open("w", encoding="utf-8") as span_file,
        section_path.open("w", encoding="utf-8") as section_file,
        memo_path.open("w", encoding="utf-8") as memo_file,
        outcome_path.open("w", encoding="utf-8") as outcome_file,
        audit_path.open("w", encoding="utf-8") as audit_file,
    ):
        for memo_index, memo_id in enumerate(memo_ids, start=1):
            memo = load_memo_record(workspace, memo_id)
            review = load_review(workspace, memo_id)
            sections = load_sections(workspace, memo_id)
            tags = load_tags(workspace, memo_id)
            evidence = load_evidence(workspace, memo_id)
            facilities = load_facilities(workspace, memo_id)
            outcome_summaries = load_outcome_summaries(workspace, memo_id)
            outcome_events = load_outcome_events(workspace, memo_id)
            foreseeability = load_foreseeability_assessments(workspace, memo_id)
            table_metrics = load_table_metrics(workspace, memo_id)
            page_quality = read_json(memo_dir(workspace, memo_id) / "extraction" / "page_quality.json", [])
            extraction_warnings = read_json(memo_dir(workspace, memo_id) / "extraction" / "ocr_warnings.json", [])
            schema_snapshot = read_json(memo_dir(workspace, memo_id) / "schema" / "schema_snapshot.json", {})
            export_memo_id = f"export_memo_{memo_index:06}"
            section_ids = _record_export_ids(sections, "section_id", "export_section", memo_index)
            facility_ids = _record_export_ids(facilities, "facility_id", "export_facility", memo_index)
            evidence_ids = _record_export_ids(evidence, "evidence_id", "export_evidence", memo_index)
            tag_ids = _record_export_ids(tags, "tag_record_id", "export_tag", memo_index)
            summary_ids = _record_export_ids(outcome_summaries, "outcome_summary_id", "export_outcome_summary", memo_index)
            event_ids = _record_export_ids(outcome_events, "outcome_event_id", "export_outcome_event", memo_index)
            foreseeability_ids = _record_export_ids(foreseeability, "foreseeability_id", "export_foreseeability", memo_index)

            memo_evidence = [item for item in evidence if item.get("evidence_type", "memo_evidence") == "memo_evidence"]
            evidence_lookup = _evidence_by_id(memo_evidence)
            quality_by_page = {int(item.get("page_number")): item for item in page_quality if item.get("page_number")}
            facilities_by_id = {facility.get("facility_id"): facility for facility in facilities}
            memo_text = "\n\n".join(
                f"[{section.get('canonical_section_name')} / {section.get('original_header')}]\n{section.get('text', '')}"
                for section in sections
            )
            memo_mapping = {
                "export_memo_id": export_memo_id,
                "memo_id": memo_id,
                "customer_id": memo.get("customer_id", ""),
                "source_file_name": memo.get("source_file_name", ""),
                "source_hash": memo.get("source_hash", ""),
                "sections": section_ids,
                "facilities": facility_ids,
                "evidence": evidence_ids,
                "tags": tag_ids,
                "outcome_summaries": summary_ids,
                "outcome_events": event_ids,
                "foreseeability_assessments": foreseeability_ids,
            }
            manifest["memos"].append(memo_mapping)

            for evidence_record in memo_evidence:
                evidence_id = str(evidence_record.get("evidence_id", ""))
                linked_tags = _tags_for_evidence(tags, evidence_id)
                if not linked_tags:
                    continue
                section = next((item for item in sections if item.get("section_id") == evidence_record.get("section_id")), {})
                response = {
                    "schema_version": SCHEMA_VERSION,
                    "export_memo_id": export_memo_id,
                    "evidence": _training_evidence(evidence_record, evidence_ids, section_ids, facility_ids),
                    "tags": [_training_tag(tag, tag_ids, facility_ids, section_ids, evidence_ids) for tag in linked_tags],
                    "facilities": [
                        _training_facility(facilities_by_id[facility_id], facility_ids)
                        for facility_id in evidence_record.get("facility_ids", [])
                        if facility_id in facilities_by_id
                    ],
                }
                span_file.write(
                    _json_dumps(
                        {
                            "instruction": "Map this cited credit memo evidence span to the reviewed underwriting tags it supports.",
                            "context": (
                                f"Export memo ID: {export_memo_id}\n"
                                f"Canonical section: {section.get('canonical_section_name', '')}\n"
                                f"Original heading: {section.get('original_header', '')}\n"
                                f"Source location: {evidence_record.get('source_location', '')}\n\n"
                                f"Evidence text:\n{evidence_record.get('selected_text', '')}"
                            ),
                            "response": _json_dumps(response),
                        }
                    )
                    + "\n"
                )

            for section in sections:
                section_tags = _tags_for_section(tags, section.get("section_id"))
                section_evidence_ids = {
                    evidence_id
                    for tag in section_tags
                    for evidence_id in tag.get("evidence_ids", [])
                    if evidence_id in evidence_lookup
                }
                response = {
                    "schema_version": SCHEMA_VERSION,
                    "export_memo_id": export_memo_id,
                    "section": _training_section(section, section_ids),
                    "extraction_quality": {
                        "page_start": section.get("page_start"),
                        "page_end": section.get("page_end"),
                        "page_quality": [
                            _strip_keys(quality_by_page[page_number], TRAINING_IDENTITY_KEYS)
                            for page_number in range(int(section.get("page_start", 1)), int(section.get("page_end", 1)) + 1)
                            if page_number in quality_by_page
                        ],
                    },
                    "tags": [_training_tag(tag, tag_ids, facility_ids, section_ids, evidence_ids) for tag in section_tags],
                    "facilities": [
                        _training_facility(facility, facility_ids)
                        for facility in facilities
                        if facility.get("source_section_id") == section.get("section_id")
                        or any(tag.get("facility_id") == facility.get("facility_id") for tag in section_tags)
                    ],
                    "evidence": [
                        _training_evidence(evidence_lookup[evidence_id], evidence_ids, section_ids, facility_ids)
                        for evidence_id in sorted(section_evidence_ids)
                    ],
                }
                quality_context = "; ".join(
                    f"p.{page_number}: {quality_by_page[page_number].get('status')}"
                    for page_number in range(int(section.get("page_start", 1)), int(section.get("page_end", 1)) + 1)
                    if page_number in quality_by_page
                )
                context = (
                    f"Export memo ID: {export_memo_id}\n"
                    f"Memo type: {memo.get('memo_type', '')}\n"
                    f"Facility type: {memo.get('facility_type', '')}\n"
                    f"Original heading: {section.get('original_header', '')}\n"
                    f"Canonical section: {section.get('canonical_section_name', '')}\n"
                    f"Page range: {section.get('page_start')} to {section.get('page_end')}\n\n"
                    f"Text quality: {quality_context or 'Not available'}\n\n"
                    f"Section text:\n{section.get('text', '')}"
                )
                section_file.write(
                    _json_dumps(
                        {
                            "instruction": "Extract CRAIG underwriting tags for this credit memo section using the canonical section mapping and cited evidence.",
                            "context": context,
                            "response": _json_dumps(response),
                        }
                    )
                    + "\n"
                )

            memo_response = {
                "schema_version": SCHEMA_VERSION,
                "training_lane": "as_of_credit_review",
                "export_memo_id": export_memo_id,
                "memo": _training_memo(memo, export_memo_id),
                "dataset_status": review.get("status", ""),
                "sections": [_training_section(section, section_ids) for section in sections],
                "facilities": [_training_facility(facility, facility_ids) for facility in facilities],
                "table_metrics": _strip_keys(table_metrics, TRAINING_IDENTITY_KEYS),
                "tags": [_training_tag(tag, tag_ids, facility_ids, section_ids, evidence_ids) for tag in tags],
                "evidence": [_training_evidence(item, evidence_ids, section_ids, facility_ids) for item in memo_evidence],
                "extraction_quality": {
                    "page_quality": _strip_keys(page_quality, TRAINING_IDENTITY_KEYS),
                    "warnings": _strip_keys(extraction_warnings, TRAINING_IDENTITY_KEYS),
                },
                "schema_snapshot": {
                    "schema_version": schema_snapshot.get("schema_version"),
                    "section_count": len(schema_snapshot.get("sections", [])) if schema_snapshot else 0,
                    "tag_count": len(schema_snapshot.get("tags", [])) if schema_snapshot else 0,
                },
            }
            memo_file.write(
                _json_dumps(
                        {
                            "instruction": "Produce a complete CRAIG RCO-assistive as-of memo review using only the credit memo text and reviewed memo evidence. Do not use future performance outcomes.",
                        "context": (
                            f"Export memo ID: {export_memo_id}\n"
                            f"Memo type: {memo.get('memo_type', '')}\n"
                            f"Facility type: {memo.get('facility_type', '')}\n"
                            f"Confirmed facilities: {len(facilities)}\n"
                            f"Text quality pages: {len(page_quality)}\n\n"
                            f"Memo text:\n{memo_text}"
                        ),
                        "response": _json_dumps(memo_response),
                    }
                )
                + "\n"
            )

            if outcome_summaries or outcome_events or foreseeability:
                outcome_response = {
                    "schema_version": SCHEMA_VERSION,
                    "training_lane": "outcome_aware",
                    "export_memo_id": export_memo_id,
                    "memo": _training_memo(memo, export_memo_id),
                    "facilities": [_training_facility(facility, facility_ids) for facility in facilities],
                    "outcome_summaries": [
                        _training_outcome_summary(summary, summary_ids, facility_ids, event_ids)
                        for summary in outcome_summaries
                    ],
                    "outcome_events": [_training_outcome_event(event, facility_ids, event_ids) for event in outcome_events],
                    "foreseeability_assessments": [
                        _training_foreseeability(assessment, foreseeability_ids, facility_ids, event_ids, evidence_ids)
                        for assessment in foreseeability
                    ],
                    "as_of_tags": [_training_tag(tag, tag_ids, facility_ids, section_ids, evidence_ids) for tag in tags],
                    "linked_memo_evidence": [
                        _training_evidence(item, evidence_ids, section_ids, facility_ids)
                        for item in memo_evidence
                        if any(item.get("evidence_id") in assessment.get("memo_evidence_ids", []) for assessment in foreseeability)
                    ],
                }
                outcome_file.write(
                    _json_dumps(
                        {
                            "instruction": "Produce the outcome-aware training labels for this credit memo, including outcome availability, observed adverse events, primary adverse outcome, and foreseeability based on memo evidence.",
                            "context": (
                                f"Export memo ID: {export_memo_id}\n"
                                f"Memo type: {memo.get('memo_type', '')}\n"
                                f"Facility type: {memo.get('facility_type', '')}\n"
                                "Training lane: outcome-aware. These labels use observed performance outcomes and must remain separate from as-of review training.\n\n"
                                f"Memo text:\n{memo_text}"
                            ),
                            "response": _json_dumps(_strip_keys(outcome_response, TRAINING_IDENTITY_KEYS)),
                        }
                    )
                    + "\n"
                )

            audit_file.write(
                _json_dumps(
                    {
                        "export_memo_id": export_memo_id,
                        "memo_id": memo_id,
                        "schema_version": SCHEMA_VERSION,
                        "schema_hash": memo.get("schema_hash", ""),
                        "source_hash": memo.get("source_hash", ""),
                        "source_file_name": memo.get("source_file_name", ""),
                        "review_status": review.get("status"),
                        "include_legacy_approved": include_legacy_approved,
                        "export_id_mapping": memo_mapping,
                        "exported_at": datetime.now(UTC).isoformat(),
                    }
                )
                + "\n"
            )

    manifest_path.write_text(_json_dumps(manifest), encoding="utf-8")

    for path in [span_path, section_path, memo_path, outcome_path, audit_path, manifest_path]:
        sync_path_to_remote(workspace, path)
    return {
        "spans": span_path,
        "sections": section_path,
        "memos": memo_path,
        "outcomes": outcome_path,
        "audit": audit_path,
        "manifest": manifest_path,
    }


def export_memo_bundle(workspace: Path, memo_id: str) -> Path:
    export_dir = memo_dir(workspace, memo_id) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    path = export_dir / f"{memo_id}_bundle_{_now_stamp()}.json"
    payload = {
        "memo": load_memo_record(workspace, memo_id),
        "review": load_review(workspace, memo_id),
        "sections": load_sections(workspace, memo_id),
        "facilities": load_facilities(workspace, memo_id),
        "tags": load_tags(workspace, memo_id),
        "evidence": load_evidence(workspace, memo_id),
        "outcome_summaries": load_outcome_summaries(workspace, memo_id),
        "outcome_events": load_outcome_events(workspace, memo_id),
        "foreseeability_assessments": load_foreseeability_assessments(workspace, memo_id),
        "table_metrics": load_table_metrics(workspace, memo_id),
        "page_quality": read_json(memo_dir(workspace, memo_id) / "extraction" / "page_quality.json", []),
        "layout_blocks": read_json(memo_dir(workspace, memo_id) / "extraction" / "layout_blocks.json", []),
        "section_candidates": read_json(memo_dir(workspace, memo_id) / "sections" / "section_candidates.json", []),
        "extraction_warnings": read_json(memo_dir(workspace, memo_id) / "extraction" / "ocr_warnings.json", []),
        "schema_snapshot": read_json(memo_dir(workspace, memo_id) / "schema" / "schema_snapshot.json", {}),
        "audit_events": load_audit_events(workspace, memo_id),
    }
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    sync_path_to_remote(workspace, path)
    return path
